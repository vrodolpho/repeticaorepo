import openpyxl

book = openpyxl.load_workbook('Planilha de Compras.xlsx')
frutas_page =  book['Frutas']
rows = frutas_page.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3)
print(rows)

frutas = []
quantidades = []
for a,b,c in rows:
  frutas.append(a.value)
  quantidades.append(b.value)

print(frutas)
print(quantidades)

columns = frutas_page.iter_cols(min_row=2, max_row=5, min_col=1, max_col=3)

for col in columns:
  print(col)